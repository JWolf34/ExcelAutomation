from openpyxl import Workbook, load_workbook


def start(filepath):
    wb = load_workbook(filepath)
    ws1 = wb.active

    #ws.move_range("G2:H3", rows=2, cols=4)

    ws2 = wb.create_sheet("12.19.19")
    
    for i in range(1, 100):
        for j in range(7, 14):
            ws2.cell(row=i, column=j-6).value = ws1.cell(row=i, column=j).value

    for i in range(4):
        ws2.insert_rows(1)

    wb.save(filepath)








if __name__ == "__main__":
    filepath = "../resources/Example.xlsx"
    start(filepath)